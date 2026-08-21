"""Dashboard ciudadano multipanel, autónomo e interactivo con Plotly.js."""

from __future__ import annotations

import json
import re
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.offline import get_plotlyjs
from plotly.utils import PlotlyJSONEncoder
from scipy.stats import linregress

from src.dashboard_utils import (
    COLORS,
    LOCALITY_NAMES,
    load_dataset,
    load_geography,
    locality_key,
    normalize_locality,
    project_root,
)


CRIME_FIELDS = {
    "CMH25CONT": "Homicidios",
    "CMLP25CONT": "Lesiones personales",
    "CMHP25CONT": "Hurto a personas",
    "CMHR25CONT": "Hurto a residencias",
    "CMHA25CONT": "Hurto de automóviles",
    "CMHB25CONT": "Hurto de bicicletas",
    "CMHM25CONT": "Hurto de motocicletas",
    "CMVI25CONT": "Violencia intrafamiliar",
}


def _clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def prepare_interactive_support(
    root: Path | None = None,
    analytical_df: pd.DataFrame | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, list[Path]]:
    """Extrae solo el detalle adicional requerido por los controles interactivos."""
    root = project_root(root)
    df = load_dataset(root) if analytical_df is None else analytical_df.copy()
    allowed = set(df["locality"].map(locality_key))

    crime_path = root / "data" / "raw" / "delitos_alto_impacto.geojson"
    with crime_path.open(encoding="utf-8-sig") as source:
        crime_geojson = json.load(source)
    detail_rows: list[dict] = []
    for feature in crime_geojson["features"]:
        properties = feature["properties"]
        raw_key = locality_key(properties.get("CMNOMLOCAL"))
        if raw_key not in LOCALITY_NAMES:
            continue
        locality = normalize_locality(properties["CMNOMLOCAL"])
        key = locality_key(locality)
        if key not in allowed:
            continue
        for field, label in CRIME_FIELDS.items():
            detail_rows.append(
                {
                    "locality": locality,
                    "locality_key": key,
                    "crime_type": label,
                    "official_count": float(properties.get(field) or 0),
                }
            )
    crime_detail = pd.DataFrame(detail_rows)
    if crime_detail.empty or crime_detail["locality"].nunique() != len(df):
        raise ValueError("No fue posible obtener el detalle de delitos para todas las localidades.")
    calculated_totals = crime_detail.groupby("locality")["official_count"].sum().sort_index()
    expected_totals = df.set_index("locality")["official_crimes_2025"].sort_index()
    if not np.allclose(calculated_totals, expected_totals):
        raise ValueError("El detalle de delitos no coincide con el total procesado existente.")

    from openpyxl import load_workbook

    survey_path = root / "data" / "raw" / "encuesta_distrital_percepcion_2025.xlsx"
    workbook = load_workbook(survey_path, read_only=True, data_only=True, keep_links=False)
    worksheet = workbook["Cuadro 16A"]
    rows = []
    for values in worksheet.iter_rows(min_row=15, max_col=11, values_only=True):
        raw_key = locality_key(values[0])
        if values[0] is None or raw_key not in LOCALITY_NAMES:
            continue
        locality = normalize_locality(values[0])
        if locality_key(locality) not in allowed:
            continue
        rows.append(
            {
                "locality": locality,
                "category": _clean(values[2]),
                "people": float(values[3] or 0),
                "households": float(values[7] or 0),
            }
        )
    workbook.close()
    survey = pd.DataFrame(rows)
    affirmative = (
        survey[survey["category"].eq("Sí")]
        .groupby("locality", as_index=False)[["people", "households"]]
        .sum()
        .rename(columns={"people": "reported_people", "households": "reported_households"})
    )
    totals = (
        survey[survey["category"].eq("Total")]
        .groupby("locality", as_index=False)["households"]
        .sum()
        .rename(columns={"households": "total_household_responses"})
    )
    reporting = affirmative.merge(totals, on="locality", how="inner")
    reporting["report_share_pct"] = 100 * reporting["reported_households"] / reporting["total_household_responses"]
    reporting["locality_key"] = reporting["locality"].map(locality_key)
    if reporting["locality"].nunique() != len(df) or reporting.isna().any().any():
        raise ValueError("La encuesta no produjo información completa de denuncia por localidad.")

    processed = root / "data" / "processed"
    crime_output = processed / "dashboard_crime_detail.csv"
    reporting_output = processed / "dashboard_reporting_summary.csv"
    crime_detail.to_csv(crime_output, index=False, encoding="utf-8")
    reporting.to_csv(reporting_output, index=False, encoding="utf-8")
    return crime_detail, reporting, [crime_output, reporting_output]


def _figure_layout(title: str, height: int = 650) -> dict:
    return {
        "title": {"text": title, "x": 0, "xanchor": "left", "font": {"size": 28}},
        "height": height,
        "paper_bgcolor": "white",
        "plot_bgcolor": "white",
        "font": {"family": "Arial", "size": 18, "color": COLORS["navy"]},
        "margin": {"l": 90, "r": 45, "t": 85, "b": 75},
        "showlegend": False,
    }


def _choropleth(
    df: pd.DataFrame,
    geojson: dict,
    variable: str,
    title: str,
    hover_label: str,
    colorscale: list,
    suffix: str = "",
    zmid: float | None = None,
    extra_customdata: list[str] | None = None,
) -> go.Figure:
    map_df = df.copy()
    map_df["locality_key"] = map_df["locality"].map(locality_key)
    custom_columns = ["locality", variable, *(extra_customdata or [])]
    trace = go.Choropleth(
        geojson=geojson,
        locations=map_df["locality_key"],
        z=map_df[variable],
        zmid=zmid,
        featureidkey="properties.locality_key",
        customdata=map_df[custom_columns],
        colorscale=colorscale,
        marker_line_color="white",
        marker_line_width=1.2,
        colorbar={"thickness": 14, "tickfont": {"size": 16}},
        hovertemplate=f"<b>%{{customdata[0]}}</b><br>{hover_label}: %{{customdata[1]:,.1f}}{suffix}<extra></extra>",
    )
    fig = go.Figure(trace)
    fig.update_geos(fitbounds="locations", visible=False)
    fig.update_layout(**_figure_layout(title))
    fig.update_layout(margin={"l": 10, "r": 45, "t": 85, "b": 20})
    return fig


def create_multipanel_figures(
    df: pd.DataFrame,
    geojson: dict,
    crime_detail: pd.DataFrame,
    reporting: pd.DataFrame,
) -> tuple[dict[str, go.Figure], pd.DataFrame, dict]:
    """Construye todas las vistas Plotly usadas por el HTML navegable."""
    enriched = df.merge(reporting[["locality", "report_share_pct"]], on="locality", how="left")
    enriched["official_crime_rate_100k"] = 100000 * enriched["official_crimes_2025"] / enriched["population_2025"]
    disposition_z = (
        enriched["disposicion_inadecuada_pct"] - enriched["disposicion_inadecuada_pct"].mean()
    ) / enriched["disposicion_inadecuada_pct"].std(ddof=0)
    gap_z = (
        enriched["perception_excess_index"] - enriched["perception_excess_index"].mean()
    ) / enriched["perception_excess_index"].std(ddof=0)
    enriched["relationship_contribution"] = disposition_z * gap_z

    x = enriched["disposicion_inadecuada_pct"].to_numpy()
    y = enriched["perception_excess_index"].to_numpy()
    regression = linregress(x, y)
    xline = np.linspace(x.min(), x.max(), 100)
    relationship = go.Figure()
    relationship.add_trace(
        go.Scatter(
            x=x,
            y=y,
            mode="markers",
            marker={"size": 18, "color": COLORS["blue"], "line": {"width": 2, "color": "white"}},
            customdata=enriched[["locality"]],
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>Disposición inadecuada: %{x:.1f} %"
                "<br>Brecha de percepción: %{y:.2f} desviaciones estándar<extra></extra>"
            ),
            showlegend=False,
        )
    )
    relationship.add_trace(
        go.Scatter(
            x=xline,
            y=regression.intercept + regression.slope * xline,
            mode="lines",
            line={"width": 5, "color": COLORS["orange"]},
            hoverinfo="skip",
            showlegend=False,
        )
    )
    relationship.update_layout(**_figure_layout("Disposición inadecuada y brecha de percepción", 700))
    relationship.update_xaxes(title="Hogares con disposición inadecuada (%)", gridcolor=COLORS["grid"])
    relationship.update_yaxes(title="Brecha de percepción (desviaciones estándar)", gridcolor=COLORS["grid"])

    home = go.Figure(relationship)
    home.update_layout(
        title={"text": "Dos situaciones que tienden a coincidir", "x": 0, "font": {"size": 32}},
        height=560,
        margin={"l": 90, "r": 35, "t": 85, "b": 75},
    )

    official_map = _choropleth(
        enriched,
        geojson,
        "official_crimes_2025",
        "Crímenes de alto impacto registrados",
        "Crímenes registrados",
        [[0, "#E8F4FA"], [0.5, "#4FA3C7"], [1, "#0B4F6C"]],
    )
    official_rate_map = _choropleth(
        enriched,
        geojson,
        "official_crime_rate_100k",
        "Crímenes registrados por cada 100.000 habitantes",
        "Crímenes por cada 100.000 habitantes",
        [[0, "#EDF7F7"], [0.5, "#5CC8C5"], [1, "#006D77"]],
    )
    estimated_rate_map = _choropleth(
        enriched,
        geojson,
        "crime_rate_100k",
        "Crímenes estimados al incluir la no denuncia",
        "Crímenes estimados por cada 100.000 habitantes",
        [[0, "#FFF4E6"], [0.5, "#F4A261"], [1, "#B54708"]],
        extra_customdata=["report_share_pct"],
    )

    disposition = enriched.sort_values("disposicion_inadecuada_pct")
    disposition_bar = go.Figure(
        go.Bar(
            x=disposition["disposicion_inadecuada_pct"],
            y=disposition["locality"],
            orientation="h",
            marker_color=COLORS["orange"],
            customdata=disposition[["locality", "disposicion_inadecuada_pct"]],
            hovertemplate="<b>%{customdata[0]}</b><br>Disposición inadecuada: %{x:.1f} %<extra></extra>",
        )
    )
    disposition_bar.update_layout(**_figure_layout("Disposición inadecuada de basura por localidad", 820))
    disposition_bar.update_layout(margin={"l": 205, "r": 45, "t": 85, "b": 75})
    disposition_bar.update_xaxes(title="Hogares con disposición inadecuada (%)", gridcolor=COLORS["grid"])
    disposition_bar.update_yaxes(title="", tickfont={"size": 17})

    report_order = reporting.sort_values("reported_people")
    reporting_chart = go.Figure()
    reporting_chart.add_trace(
        go.Bar(
            x=report_order["reported_people"], y=report_order["locality"], orientation="h",
            name="Personas", marker_color=COLORS["blue"],
            hovertemplate="<b>%{y}</b><br>Personas: %{x:,.0f}<extra></extra>",
        )
    )
    reporting_chart.add_trace(
        go.Bar(
            x=report_order["reported_households"], y=report_order["locality"], orientation="h",
            name="Hogares", marker_color=COLORS["teal"],
            hovertemplate="<b>%{y}</b><br>Hogares: %{x:,.0f}<extra></extra>",
        )
    )
    reporting_chart.update_layout(**_figure_layout("Personas y hogares que reportaron haber denunciado", 850))
    reporting_chart.update_layout(
        barmode="group",
        showlegend=True,
        legend={"orientation": "h", "y": 1.07, "x": 0.34},
        margin={"l": 215, "r": 45, "t": 105, "b": 75},
    )
    reporting_chart.update_xaxes(title="Respuestas afirmativas de denuncia", gridcolor=COLORS["grid"])
    reporting_chart.update_yaxes(title="", tickfont={"size": 17})

    insecurity_map = _choropleth(
        enriched,
        geojson,
        "insecurity_noche_pct",
        "¿Qué tan seguro se siente caminando solo por su barrio de noche?",
        "Personas que se sienten inseguras",
        [[0, "#E8F4FA"], [0.5, "#4FA3C7"], [1, "#0B4F6C"]],
        suffix=" %",
    )
    gap_map = _choropleth(
        enriched,
        geojson,
        "perception_excess_index",
        "Brecha entre percepción de inseguridad y delitos estimados",
        "Brecha de percepción",
        [[0, "#1677A8"], [0.5, "#F7F8FA"], [1, "#C84C61"]],
        suffix=" desviaciones estándar",
        zmid=0,
    )
    contribution_map = _choropleth(
        enriched,
        geojson,
        "relationship_contribution",
        "Aporte de cada localidad a la relación observada",
        "Aporte local",
        [[0, "#1677A8"], [0.5, "#F7F8FA"], [1, "#F28E2B"]],
        zmid=0,
    )

    first_locality = str(enriched.nlargest(1, "official_crimes_2025").iloc[0]["locality"])
    detail_payload = {}
    for locality, group in crime_detail.groupby("locality"):
        ordered = group.sort_values("official_count")
        detail_payload[locality_key(locality)] = {
            "locality": locality,
            "labels": ordered["crime_type"].tolist(),
            "values": ordered["official_count"].round(0).astype(int).tolist(),
        }
    figures = {
        "home-hook": home,
        "official-map": official_map,
        "official-rate-map": official_rate_map,
        "estimated-rate-map": estimated_rate_map,
        "disposition-bar": disposition_bar,
        "reporting-chart": reporting_chart,
        "insecurity-map": insecurity_map,
        "gap-map": gap_map,
        "relationship-scatter": relationship,
        "contribution-map": contribution_map,
    }
    return figures, enriched, {"crime_detail": detail_payload, "first_locality": locality_key(first_locality)}


def _json(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, cls=PlotlyJSONEncoder).replace("</", "<\\/")


def export_multipanel_dashboard(
    df: pd.DataFrame,
    geojson: dict,
    crime_detail: pd.DataFrame,
    reporting: pd.DataFrame,
    root: Path | None = None,
) -> Path:
    """Exporta un único HTML con navegación y eventos de selección enlazados."""
    root = project_root(root)
    figures, enriched, payload = create_multipanel_figures(df, geojson, crime_detail, reporting)
    for figure in figures.values():
        figure.layout.template = None
    figure_json = {name: figure.to_plotly_json() for name, figure in figures.items()}
    locality_values = {
        locality_key(row.locality): {
            "locality": row.locality,
            "official_rate": float(row.official_crime_rate_100k),
            "estimated_rate": float(row.crime_rate_100k),
            "report_share": float(row.report_share_pct),
            "disposition": float(row.disposicion_inadecuada_pct),
        }
        for row in enriched.itertuples()
    }
    html = f"""<!doctype html>
<html lang="es"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Basura, percepción y crimen en Bogotá D.C.</title>
<style>
:root{{--navy:#14213D;--blue:#1677A8;--teal:#00A6A6;--orange:#F28E2B;--paper:#F4F6F8;--grid:#E1E7ED;}}
*{{box-sizing:border-box}} body{{margin:0;background:var(--paper);font-family:Arial,sans-serif;color:var(--navy);font-size:20px}}
.shell{{max-width:1600px;margin:auto;padding:34px 44px 70px}} h1{{font-size:48px;line-height:1.05;margin:0 0 10px}} h2{{font-size:38px;margin:0 0 12px}} h3{{font-size:27px;margin:0 0 10px}} p{{line-height:1.45}}
.eyebrow{{color:var(--teal);font-weight:700;margin-bottom:28px}} .panel{{display:none}} .panel.active{{display:block}}
.home-grid,.two-cols{{display:grid;grid-template-columns:minmax(0,1.55fr) minmax(320px,.75fr);gap:28px;align-items:stretch}}
.plot-card,.card,.note,.selection{{background:white;border:1px solid var(--grid);border-radius:4px;padding:20px}}
.plot{{width:100%;min-height:520px}} .nav-stack{{display:flex;flex-direction:column;gap:15px;justify-content:center}}
button{{appearance:none;border:1px solid var(--navy);background:white;color:var(--navy);padding:20px 24px;text-align:left;font:700 20px Arial;border-radius:6px;cursor:pointer;transition:.15s}}
button:hover,button:focus{{background:var(--navy);color:white;transform:translateY(-1px)}} .back{{margin:0 0 26px;padding:12px 18px;font-size:18px}}
.section-head{{margin-bottom:24px}} .section-block{{margin:30px 0 46px}} .wide{{width:100%}}
.selection{{font-size:23px;font-weight:700;margin-top:12px;border-left:7px solid var(--teal)}}
.note{{font-size:22px;border-left:7px solid var(--orange);margin-top:20px}} .steps{{display:grid;grid-template-columns:repeat(3,1fr);gap:18px;margin:25px 0}}
.step{{background:white;border-top:7px solid var(--teal);padding:22px;border-radius:4px}} .step b{{font-size:24px}}
.small{{font-size:16px;color:#5F6B7A}} .spacer{{height:12px}}
@media(max-width:900px){{.shell{{padding:24px 16px}}h1{{font-size:36px}}h2{{font-size:31px}}.home-grid,.two-cols,.steps{{grid-template-columns:1fr}}button{{font-size:18px}}}}
</style><script>{get_plotlyjs().replace('</', '<\\/')}</script></head>
<body><main class="shell">
<section id="home" class="panel active"><div class="section-head"><h1>Basura, percepción y crimen<br>en Bogotá D.C.</h1><div class="eyebrow">19 localidades · información territorial 2025</div></div>
<div class="home-grid"><div class="plot-card"><div id="home-hook" class="plot"></div></div><nav class="nav-stack" aria-label="Paneles temáticos">
<button data-panel="crime">1 — Crímenes de alto impacto</button><button data-panel="waste">2 — Disposición inadecuada de basura</button>
<button data-panel="perception">3 — Percepción frente a la tasa de crímenes</button><button data-panel="relationship">4 — Brecha de percepción y disposición inadecuada</button>
</nav></div></section>

<section id="crime" class="panel"><button class="back" data-panel="home">← Volver al inicio</button><h2>Crímenes de alto impacto</h2>
<div class="two-cols section-block"><div class="plot-card"><div id="official-map" class="plot"></div></div><div class="plot-card"><div id="crime-breakdown" class="plot"></div></div></div>
<div class="section-block plot-card"><div id="official-rate-map" class="plot"></div><div id="official-rate-selection" class="selection">Seleccione una localidad en el mapa.</div></div>
<div class="section-block plot-card"><div id="estimated-rate-map" class="plot"></div><div id="estimated-rate-selection" class="selection">Seleccione una localidad para consultar su porcentaje de denuncia.</div></div></section>

<section id="waste" class="panel"><button class="back" data-panel="home">← Volver al inicio</button><h2>Disposición inadecuada de basura</h2>
<div class="section-block plot-card"><div id="disposition-bar" class="plot"></div><div id="disposition-selection" class="selection">Seleccione una barra para consultar el porcentaje.</div></div>
<div class="note">Esta variable se comparará con un índice construido por localidad sobre la brecha de percepción.</div></section>

<section id="perception" class="panel"><button class="back" data-panel="home">← Volver al inicio</button><h2>Percepción frente a la tasa de crímenes</h2>
<div class="steps"><div class="step"><b>1. Medimos</b><p>La inseguridad al caminar de noche y los crímenes estimados por cada 100.000 habitantes.</p></div>
<div class="step"><b>2. Igualamos la escala</b><p>Cada dato se expresa según su distancia frente al promedio de las localidades.</p></div>
<div class="step"><b>3. Calculamos la brecha</b><p>Percepción estandarizada menos tasa de crímenes estandarizada.</p></div></div>
<div class="section-block plot-card"><div id="reporting-chart" class="plot"></div><p class="small">Las cifras suman respuestas por tipo de delito; una persona puede aparecer en más de una categoría.</p></div>
<div class="section-block plot-card"><h2>¿Qué tan seguro se siente usted caminando solo por su barrio de noche?</h2><div id="insecurity-map" class="plot"></div></div>
<div class="section-block plot-card"><div id="gap-map" class="plot"></div><div class="note"><b>Cómo leerlo:</b> 0 representa la brecha promedio. +1 significa una desviación estándar por encima; −1, una por debajo.</div></div></section>

<section id="relationship" class="panel"><button class="back" data-panel="home">← Volver al inicio</button><h2>Brecha de percepción y disposición inadecuada</h2>
<div class="section-block plot-card"><div id="relationship-scatter" class="plot"></div></div>
<div class="section-block plot-card"><div id="contribution-map" class="plot"></div><p class="small">Los valores positivos muestran localidades que refuerzan la relación general; los negativos se apartan de ella.</p></div></section>
</main><script>
const FIGURES={_json(figure_json)}, DETAIL={_json(payload['crime_detail'])}, VALUES={_json(locality_values)}, FIRST={_json(payload['first_locality'])};
const CONFIG={{responsive:true,displaylogo:false,scrollZoom:true,modeBarButtonsToRemove:['lasso2d']}};
const INITIAL_PANEL=location.hash.slice(1);if(INITIAL_PANEL)history.replaceState(null,'',location.pathname);
Object.entries(FIGURES).forEach(([id,fig])=>Plotly.newPlot(id,fig.data,fig.layout,CONFIG));
function showPanel(id){{document.querySelectorAll('.panel').forEach(x=>x.classList.remove('active'));document.getElementById(id).classList.add('active');if(location.hash.slice(1)!==id)history.replaceState(null,'',id==='home'?location.pathname:'#'+id);window.scrollTo(0,0);setTimeout(()=>document.querySelectorAll('#'+id+' .plot').forEach(x=>Plotly.Plots.resize(x)),80)}}
document.querySelectorAll('[data-panel]').forEach(b=>b.addEventListener('click',()=>showPanel(b.dataset.panel)));
function updateBreakdown(key){{const d=DETAIL[key];if(!d)return;const method=document.getElementById('crime-breakdown').data?Plotly.react:Plotly.newPlot;method('crime-breakdown',[{{type:'bar',orientation:'h',x:d.values,y:d.labels,marker:{{color:'#1677A8'}},hovertemplate:'%{{y}}: %{{x:,.0f}}<extra></extra>'}}],{{title:{{text:'Delitos registrados en '+d.locality,x:0,font:{{size:27}}}},height:650,paper_bgcolor:'white',plot_bgcolor:'white',font:{{family:'Arial',size:17,color:'#14213D'}},margin:{{l:190,r:30,t:85,b:60}},xaxis:{{title:'Número de delitos',gridcolor:'#E1E7ED'}}}},CONFIG)}}
updateBreakdown(FIRST);
document.getElementById('official-map').on('plotly_click',e=>updateBreakdown(e.points[0].location));
document.getElementById('official-rate-map').on('plotly_click',e=>{{const d=VALUES[e.points[0].location];document.getElementById('official-rate-selection').textContent=d.locality+': '+Math.round(d.official_rate).toLocaleString('es-CO')+' crímenes registrados por cada 100.000 habitantes.'}});
document.getElementById('estimated-rate-map').on('plotly_click',e=>{{const d=VALUES[e.points[0].location];document.getElementById('estimated-rate-selection').textContent=d.locality+': '+d.report_share.toLocaleString('es-CO',{{maximumFractionDigits:1}})+' % de los hogares afirmó haber denunciado.'}});
document.getElementById('disposition-bar').on('plotly_click',e=>{{const d=e.points[0].customdata;document.getElementById('disposition-selection').textContent=d[0]+': '+Number(d[1]).toLocaleString('es-CO',{{maximumFractionDigits:1}})+' % de hogares con disposición inadecuada.'}});
if(['crime','waste','perception','relationship'].includes(INITIAL_PANEL))showPanel(INITIAL_PANEL);
</script></body></html>"""
    output = root / "outputs" / "dashboard_bogota.html"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(html, encoding="utf-8")
    if output.stat().st_size <= 1_000_000:
        raise AssertionError("El HTML no contiene Plotly completo o quedó incompleto.")
    return output
